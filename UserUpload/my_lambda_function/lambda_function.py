import json
import base64
from openpyxl import load_workbook

def lambda_handler(event, context):
    try:
        # Check if 'body' is present in the event
        if 'body' in event:
            # Assuming the 'body' contains base64-encoded file data
            encoded_file = event['body']
            decoded_file = base64.b64decode(encoded_file)
            
            # Process the Excel file using openpyxl or any other library
            workbook = load_workbook(filename=decoded_file)
               # Assuming the data is in the first sheet of the Excel file
            sheet = workbook.active

            # Extract headers from the first row
            headers = [cell.value for cell in sheet[1]]

            # Extract and print data from subsequent rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                user_data = dict(zip(headers, row))
                print(user_data)
            
            return {
                'statusCode': 200,
                'body': json.dumps('File processed successfully!')
            }
        else:
            # Return an error response if 'body' is not present
            return {
                'statusCode': 400,
                'body': json.dumps('Missing request body.')
            }
    except Exception as e:
        return {
            'statusCode': 500,
            'body': json.dumps(f'Error: {str(e)}')
        }
